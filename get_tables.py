from openpyxl import load_workbook
import os
import shutil
import zipfile
def main_tables():
  shutil.rmtree("Exit")
  os.makedirs('Exit', exist_ok=True) # создание папки


  def contains_digit(s): # Проверка на число
      return any(char.isdigit() for char in s)
    
    
  main_table = load_workbook("Entrance/all.xlsx")
  sheet = main_table.active

  arr = []

  length_first_column = sheet.max_row # Получение длины первого столбца

  for number_row in range(length_first_column):
    if_number = sheet[f'A{number_row+1}'].value # Перебераем столбец A
    
    if contains_digit(str(if_number)): 
        arr.append(if_number) # Добавляем значения в массив, для получения первого значения и последнего
        
  firs_value = arr[0] 
  end_value = arr[-1]


  for number_row in range(length_first_column):
    
      if_number = sheet[f'A{number_row+1}'].value 
      if contains_digit(str(if_number)):

        if if_number == firs_value: # Проверка на первое значение
          sheet.delete_rows(idx=number_row+2, amount=len(arr)-1)
        
        
        elif if_number == end_value: # Проверка на последнее значение
            sheet.delete_rows(idx=4, amount=len(arr)-1)


        else:
            sheet.delete_rows(idx=number_row+2, amount=33-number_row)
            sheet.delete_rows(idx=4, amount=number_row-3)
          
          
        sheet.merge_cells("C5:E5") # Объединение ячеек
        sheet.merge_cells("F5:H5") # Объединение ячеек
        sheet.merge_cells("I5:K5") # Объединение ячеек
        sheet.merge_cells("I5:K5") # Объединение ячеек
        sheet.merge_cells("L5:N5") # Объединение ячеек
        sheet.merge_cells("O5:Q5") # Объединение ячеек
        sheet.merge_cells("R5:T5") # Объединение ячеек
        sheet.merge_cells("U5:W5") # Объединение ячеек
        sheet.merge_cells("X5:Z5") # Объединение ячеек
        sheet.merge_cells("AA5:AC5") # Объединение ячеек
        sheet.merge_cells("AD5:AF5") # Объединение ячеек
        sheet.merge_cells("AG5:AI5") # Объединение ячеек
        sheet.merge_cells("AJ5:AL5") # Объединение ячеек
        sheet.merge_cells("AM5:AO5") # Объединение ячеек
        sheet.merge_cells("AP5:AR5") # Объединение ячеек
        sheet.merge_cells("AS5:AU5") # Объединение ячеек
        sheet.merge_cells("AV5:AX5") # Объединение ячеек
        sheet.merge_cells("AY5:BA5") # Объединение ячеек
        sheet.merge_cells("BB5:BD5") # Объединение ячеек
        sheet.merge_cells("BE5:BG5") # Объединение ячеек



        main_table.save(f"Exit/{sheet['B4'].value}.xlsx") # Сохраняем новую таблицу
        
        
        # Заново копирует основную таблицу
        main_table = load_workbook("Entrance/all.xlsx")
        sheet = main_table.active
        
        

  # Указываем имя папки и имя ZIP-файла
  folder_name = 'Exit'
  zip_file_name = 'Exit.zip'

  # Создаем ZIP-файл
  with zipfile.ZipFile(zip_file_name, 'w', zipfile.ZIP_DEFLATED) as zipf:
      for root, dirs, files in os.walk(folder_name):
          for file in files:
              file_path = os.path.join(root, file)
              zipf.write(file_path, os.path.relpath(file_path, folder_name))


def delete_folder():
    shutil.rmtree("Entrance")
    os.makedirs('Entrance', exist_ok=True) # создание папки